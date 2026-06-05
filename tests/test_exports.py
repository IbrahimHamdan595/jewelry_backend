"""Phase 3 export tests: xlsx workbooks re-open via openpyxl; PDFs start with %PDF.

The PDF tests exercise the bilingual RTL builders + WeasyPrint (Pango/HarfBuzz),
so they require the system font/text stack from the backend Dockerfile.
"""
import io
from datetime import date
from decimal import Decimal as D
from types import SimpleNamespace

import openpyxl

from app.api.accounting import _tb_sheets
from app.api.ar import _invoice_html, _statement_html
from app.api.expenses import _bill_html
from app.api.statements import _bs_pdf, _cf_pdf, _pnl_pdf
from app.api.tax import _vat_sheets
from app.core import pdf
from app.core.xlsx import XLSX_MEDIA_TYPE, build_xlsx_bytes


# ── XLSX ──────────────────────────────────────────────────────────────────────

def test_trial_balance_xlsx_reopens():
    tb = {
        "as_of": date(2026, 6, 30), "total_base_debit": D("1000.00"), "total_base_credit": D("1000.00"),
        "accounts": [
            {"code": "1000", "name": "Cash", "base_debit": D("1000.00"), "base_credit": D("0.00"),
             "metal_by_karat": {}},
            {"code": "4000", "name": "Sales", "base_debit": D("0.00"), "base_credit": D("1000.00"),
             "metal_by_karat": {"K21": {"net_grams": D("5.000")}}},
        ],
    }
    body = build_xlsx_bytes(_tb_sheets(tb))
    assert body[:2] == b"PK"
    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["Trial Balance"]
    ws = wb["Trial Balance"]
    assert ws.cell(row=1, column=1).value == "Trial Balance as of 2026-06-30"
    # last data row is the Total
    assert ws.cell(row=ws.max_row, column=2).value == "Total"
    assert ws.cell(row=ws.max_row, column=3).value == 1000.0


def test_vat_return_xlsx_reopens():
    vat = {"year": 2026, "quarter": 2, "from": date(2026, 4, 1), "until": date(2026, 6, 30),
           "output_vat": D("300.00"), "input_vat": D("120.00"), "net_payable": D("180.00"),
           "direction": "payable",
           "transactions": [{"entry_no": "JE-1", "date": date(2026, 5, 1),
                             "source_type": "AR_INVOICE", "kind": "output", "vat": D("300.00")}]}
    body = build_xlsx_bytes(_vat_sheets(vat))
    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["VAT Return", "Transactions"]
    ws = wb["VAT Return"]
    assert ws.cell(row=2, column=1).value == "Item"            # title row 1, headers row 2
    assert ws.cell(row=3, column=1).value == "Output VAT"      # data starts row 3
    assert ws.cell(row=3, column=2).value == 300.0


# ── PDF builders ──────────────────────────────────────────────────────────────

def _stmt_data():
    return {"from": date(2026, 6, 1), "until": date(2026, 6, 30),
            "events": [
                {"date": date(2026, 6, 5), "kind": "invoice", "ref": "INV-1",
                 "debit": D("1000.00"), "credit": D("0.00"), "balance": D("1000.00")},
                {"date": date(2026, 6, 10), "kind": "receipt", "ref": "RC-1",
                 "debit": D("0.00"), "credit": D("400.00"), "balance": D("600.00")},
            ],
            "closing_balance": D("600.00")}


def test_ar_statement_pdf_english():
    html = _statement_html(_stmt_data(), "Acme Trading", "en")
    assert 'dir="ltr"' in html and "Account Statement" in html
    assert pdf.render_pdf(html)[:4] == b"%PDF"


def test_ar_statement_pdf_arabic_rtl():
    html = _statement_html(_stmt_data(), "محمد العامل", "ar")
    # RTL shell + Arabic title + Arabic customer name embedded
    assert 'dir="rtl"' in html
    assert "كشف حساب" in html and "محمد العامل" in html
    body = pdf.render_pdf(html)
    assert body[:4] == b"%PDF"
    assert len(body) > 1500  # a real rendered page, not an empty stub


def test_invoice_pdf_arabic():
    inv = SimpleNamespace(invoice_no="INV-100", invoice_date=date(2026, 6, 5), customer_id="c1",
                          subtotal=D("1000.00"), vat_amount=D("110.00"), total=D("1110.00"),
                          amount_paid=D("500.00"))
    lines = [SimpleNamespace(description="خاتم ذهب", quantity=1, unit_price=D("1000.00"), line_total=D("1000.00"))]
    html = _invoice_html(inv, lines, "محمد العامل", "ar")
    assert 'dir="rtl"' in html and "فاتورة" in html
    assert pdf.render_pdf(html)[:4] == b"%PDF"


def test_bill_pdf_english():
    bill = SimpleNamespace(bill_no="BILL-9", vendor_name="Gold Supplier", bill_date=date(2026, 6, 5),
                           subtotal=D("800.00"), vat_amount=D("0.00"), total=D("800.00"),
                           amount_paid=D("0.00"))
    lines = [SimpleNamespace(description="Rent", amount=D("800.00"))]
    html = _bill_html(bill, lines, "en")
    assert pdf.render_pdf(html)[:4] == b"%PDF"


def test_three_statements_pdf():
    pnl = {"start": date(2026, 6, 1), "end": date(2026, 6, 30),
           "revenue_lines": [{"name": "Sales", "amount": D("1000.00")}],
           "cogs_lines": [{"name": "Metal", "amount": D("600.00")}],
           "opex_lines": [{"name": "Rent", "amount": D("100.00")}],
           "other_lines": [{"name": "FX gain", "amount": D("27.17")}],
           "revenue": D("1000.00"), "cogs": D("600.00"), "gross_profit": D("400.00"),
           "operating_expenses": D("100.00"), "operating_profit": D("300.00"),
           "other_income_expense": D("27.17"), "net_profit": D("327.17")}
    bs = {"as_of": date(2026, 6, 30),
          "asset_lines": [{"name": "Cash", "amount": D("5000.00")}],
          "liability_lines": [{"name": "AP", "amount": D("1000.00")}],
          "equity_lines": [{"name": "Capital", "amount": D("4000.00")}],
          "total_assets": D("5000.00"), "total_liabilities": D("1000.00"), "total_equity": D("4000.00")}
    cf = {"start": date(2026, 6, 1), "end": date(2026, 6, 30),
          "opening_cash": D("5000.00"), "closing_cash": D("6350.00"), "net_change": D("1350.00"),
          "categories": [{"label": "Sales", "amount": D("1000.00")}]}
    for html in (_pnl_pdf(pnl, "ar"), _bs_pdf(bs, "ar"), _cf_pdf(cf, "en")):
        assert pdf.render_pdf(html)[:4] == b"%PDF"


def test_pdf_and_xlsx_response_headers():
    pdf_resp = pdf.pdf_response(_statement_html(_stmt_data(), "Acme", "en"), filename="statement-x")
    assert pdf_resp.media_type == "application/pdf"
    assert 'filename="statement-x.pdf"' in pdf_resp.headers["content-disposition"]

    from app.core.xlsx import build_xlsx_response
    xlsx_resp = build_xlsx_response(_tb_sheets({
        "as_of": date(2026, 6, 30), "total_base_debit": D("0.00"), "total_base_credit": D("0.00"),
        "accounts": []}), filename="tb-x")
    assert xlsx_resp.media_type == XLSX_MEDIA_TYPE
    assert 'filename="tb-x.xlsx"' in xlsx_resp.headers["content-disposition"]
