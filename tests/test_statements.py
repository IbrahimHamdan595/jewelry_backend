from datetime import date
from decimal import Decimal as D

import pytest
from sqlalchemy import select

from app.core import gl, statements
from app.core.coa_seed import seed_chart_of_accounts
from app.models import GLAccount, GLPeriod, PeriodStatus


async def _acct(db, system_key: str) -> str:
    return (await db.execute(
        select(GLAccount.id).where(GLAccount.system_key == system_key)
    )).scalar_one()


async def _seed(db):
    """CoA + OPEN periods for May & June 2026 + bank_accounts for cash ids."""
    await seed_chart_of_accounts(db)
    from app.core.bank import adopt_seeded_accounts
    await adopt_seeded_accounts(db)
    db.add(GLPeriod(year=2026, period_no=5, status=PeriodStatus.OPEN))
    db.add(GLPeriod(year=2026, period_no=6, status=PeriodStatus.OPEN))
    await db.flush()


async def _post(db, entry_date, lines, memo="t"):
    return await gl.post_entry(
        db, entry_date=entry_date, memo=memo, source_type="TEST",
        source_id=None, lines=lines, actor_user_id="u1",
    )


def _m(account_id, *, debit=D("0"), credit=D("0")):
    """A MONEY line (base + money set equal)."""
    return gl.GLLine(account_id=account_id, denomination="MONEY",
                     base_debit=debit, base_credit=credit,
                     money_debit=debit, money_credit=credit, currency="USD")


@pytest.mark.asyncio
async def test_income_statement_revenue_cogs_netprofit(db):
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")
    cogs = await _acct(db, "MAKING_COGS")
    inv = await _acct(db, "PRODUCT_INVENTORY")
    rent = await _acct(db, "RENT_EXPENSE")

    # Cash sale: DR cash 1000 / CR revenue 1000
    await _post(db, date(2026, 6, 5), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    # COGS: DR cogs 600 / CR inventory 600
    await _post(db, date(2026, 6, 5), [_m(cogs, debit=D("600")), _m(inv, credit=D("600"))])
    # Rent expense paid: DR rent 150 / CR cash 150
    await _post(db, date(2026, 6, 6), [_m(rent, debit=D("150")), _m(cash, credit=D("150"))])

    pnl = await statements.income_statement(db, start=date(2026, 6, 1), end=date(2026, 6, 30))
    assert pnl["revenue"] == D("1000.00")
    assert pnl["cogs"] == D("600.00")
    assert pnl["gross_profit"] == D("400.00")
    assert pnl["operating_expenses"] == D("150.00")
    assert pnl["net_profit"] == D("250.00")
    # COGS is split out from opex
    assert {l["system_key"] for l in pnl["cogs_lines"]} == {"MAKING_COGS"}
    assert {l["system_key"] for l in pnl["opex_lines"]} == {"RENT_EXPENSE"}


@pytest.mark.asyncio
async def test_balance_sheet_identity_and_metal_schedule(db):
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")
    cogs = await _acct(db, "MAKING_COGS")
    inv = await _acct(db, "PRODUCT_INVENTORY")
    metal_inv = await _acct(db, "METAL_INVENTORY")
    obe = await _acct(db, "OPENING_BALANCE_EQUITY")

    # Opening: owner funds cash 5000 (DR cash / CR opening-balance-equity)
    await _post(db, date(2026, 6, 1), [_m(cash, debit=D("5000")), _m(obe, credit=D("5000"))])
    # Opening metal inventory: 100g K21 valued 4000. Both dimensions must net to
    # zero, so the OPENING_BALANCE_EQUITY counterpart (DUAL) carries the metal too.
    await _post(db, date(2026, 6, 1), [
        gl.GLLine(account_id=metal_inv, denomination="DUAL",
                  base_debit=D("4000"), money_debit=D("4000"), currency="USD",
                  metal_debit_grams=D("100"), karat="K21"),
        gl.GLLine(account_id=obe, denomination="DUAL",
                  base_credit=D("4000"), money_credit=D("4000"), currency="USD",
                  metal_credit_grams=D("100"), karat="K21"),
    ])
    # A cash sale with COGS so current-period earnings is non-zero
    await _post(db, date(2026, 6, 10), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    await _post(db, date(2026, 6, 10), [_m(cogs, debit=D("600")), _m(inv, credit=D("600"))])

    bs = await statements.balance_sheet(db, as_of=date(2026, 6, 30))
    assert bs["balanced"] is True
    assert bs["total_assets"] == bs["total_liabilities"] + bs["total_equity"]
    # Current-period earnings line present = net income (1000 - 600 = 400)
    earnings = [l for l in bs["equity_lines"] if l["system_key"] == "CURRENT_EARNINGS"]
    assert earnings and earnings[0]["amount"] == D("400.00")
    # Metal schedule: 100g K21
    sched = {m["karat"]: m["net_grams"] for m in bs["metal_position"]}
    assert sched.get("K21") == D("100.000")
    assert bs["all_current"] is True


@pytest.mark.asyncio
async def test_cash_flow_reconciles_and_categorizes(db):
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")
    rent = await _acct(db, "RENT_EXPENSE")
    obe = await _acct(db, "OPENING_BALANCE_EQUITY")

    # Before window: owner funds 5000 (sets opening cash)
    await _post(db, date(2026, 5, 20), [_m(cash, debit=D("5000")), _m(obe, credit=D("5000"))])
    # In window: cash sale +1000 ; rent paid -150 ; owner adds +500
    await _post(db, date(2026, 6, 5), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    await _post(db, date(2026, 6, 6), [_m(rent, debit=D("150")), _m(cash, credit=D("150"))])
    await _post(db, date(2026, 6, 7), [_m(cash, debit=D("500")), _m(obe, credit=D("500"))])

    cf = await statements.cash_flow(db, start=date(2026, 6, 1), end=date(2026, 6, 30))
    assert cf["opening_cash"] == D("5000.00")
    assert cf["closing_cash"] == D("6350.00")
    assert cf["net_change"] == D("1350.00")
    assert cf["reconciles"] is True
    cats = {c["key"]: c["amount"] for c in cf["categories"]}
    assert cats["sales"] == D("1000.00")
    assert cats["expenses"] == D("-150.00")
    assert cats["owner"] == D("500.00")


@pytest.mark.asyncio
async def test_account_statement_running_balance(db):
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")
    rent = await _acct(db, "RENT_EXPENSE")
    obe = await _acct(db, "OPENING_BALANCE_EQUITY")

    await _post(db, date(2026, 5, 25), [_m(cash, debit=D("5000")), _m(obe, credit=D("5000"))])  # opening
    await _post(db, date(2026, 6, 5), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    await _post(db, date(2026, 6, 6), [_m(rent, debit=D("150")), _m(cash, credit=D("150"))])

    st = await statements.account_statement(
        db, account_id=cash, start=date(2026, 6, 1), end=date(2026, 6, 30))
    assert st["opening_balance"] == D("5000.00")   # from before the window
    assert [r["running_balance"] for r in st["rows"]] == [D("6000.00"), D("5850.00")]
    assert st["closing_balance"] == D("5850.00")
    assert st["code"] and st["name"]


@pytest.mark.asyncio
async def test_general_ledger_drilldown_reconciles_to_trial_balance(db):
    """Phase 2 GL drilldown invariant: opening + Σ(debit−credit) within the window
    equals closing, and closing reconciles to the account's trial-balance net."""
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")
    rent = await _acct(db, "RENT_EXPENSE")
    obe = await _acct(db, "OPENING_BALANCE_EQUITY")

    await _post(db, date(2026, 5, 25), [_m(cash, debit=D("5000")), _m(obe, credit=D("5000"))])  # before window
    await _post(db, date(2026, 6, 5), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    await _post(db, date(2026, 6, 6), [_m(rent, debit=D("150")), _m(cash, credit=D("150"))])

    st = await statements.account_statement(
        db, account_id=cash, start=date(2026, 6, 1), end=date(2026, 6, 30))
    movement = sum((r["debit"] - r["credit"] for r in st["rows"]), D("0"))
    assert st["opening_balance"] + movement == st["closing_balance"]

    # Closing balance must equal this account's net in the trial balance at window end.
    tb = await gl.compute_trial_balance(db, as_of=date(2026, 6, 30))
    cash_net = [a for a in tb["accounts"] if a["account_id"] == cash][0]["net_base"]
    assert cash_net == st["closing_balance"] == D("5850.00")


@pytest.mark.asyncio
async def test_general_ledger_metal_dimension_reconciles_per_karat(db):
    """The grams columns aren't decoration: opening grams + Σ(metal_debit−metal_credit)
    within the window must equal closing grams, and closing grams must tie to the
    account's per-karat net in the trial balance."""
    await _seed(db)
    metal_inv = await _acct(db, "METAL_INVENTORY")
    obe = await _acct(db, "OPENING_BALANCE_EQUITY")

    def _dual(account_id, *, dr_g=D("0"), cr_g=D("0"), base):
        return gl.GLLine(account_id=account_id, denomination="DUAL",
                         base_debit=base if dr_g else D("0"), base_credit=base if cr_g else D("0"),
                         money_debit=base if dr_g else D("0"), money_credit=base if cr_g else D("0"),
                         currency="USD", metal_debit_grams=dr_g, metal_credit_grams=cr_g, karat="K21")

    # Before window: 100g K21 opening into metal inventory.
    await _post(db, date(2026, 5, 25), [_dual(metal_inv, dr_g=D("100"), base=D("4000")),
                                        _dual(obe, cr_g=D("100"), base=D("4000"))])
    # In window: +50g, then −30g (both K21).
    await _post(db, date(2026, 6, 5), [_dual(metal_inv, dr_g=D("50"), base=D("2000")),
                                       _dual(obe, cr_g=D("50"), base=D("2000"))])
    await _post(db, date(2026, 6, 10), [_dual(metal_inv, cr_g=D("30"), base=D("1200")),
                                        _dual(obe, dr_g=D("30"), base=D("1200"))])

    st = await statements.account_statement(
        db, account_id=metal_inv, start=date(2026, 6, 1), end=date(2026, 6, 30))
    assert st["opening_grams"] == D("100.000")
    movement = sum((r["metal_debit_grams"] - r["metal_credit_grams"] for r in st["rows"]), D("0"))
    assert movement == D("20.000")
    assert st["opening_grams"] + movement == st["closing_grams"] == D("120.000")
    assert st["rows"][-1]["running_grams"] == D("120.000")

    # Ties to METAL_INVENTORY's K21 net in the trial balance.
    tb = await gl.compute_trial_balance(db, as_of=date(2026, 6, 30))
    row = [a for a in tb["accounts"] if a["account_id"] == metal_inv][0]
    assert row["metal_by_karat"]["K21"]["net_grams"] == D("120.000")


@pytest.mark.asyncio
async def test_statements_exclude_entries_after_window(db):
    await _seed(db)
    cash = await _acct(db, "CASH")
    rev = await _acct(db, "SALES_REVENUE")

    # In-window sale 1000, plus a JULY sale 9999 that must be excluded.
    await _post(db, date(2026, 6, 5), [_m(cash, debit=D("1000")), _m(rev, credit=D("1000"))])
    db.add(GLPeriod(year=2026, period_no=7, status=PeriodStatus.OPEN))
    await db.flush()
    await _post(db, date(2026, 7, 5), [_m(cash, debit=D("9999")), _m(rev, credit=D("9999"))])

    pnl = await statements.income_statement(db, start=date(2026, 6, 1), end=date(2026, 6, 30))
    assert pnl["revenue"] == D("1000.00")
    bs = await statements.balance_sheet(db, as_of=date(2026, 6, 30))
    earnings = [l for l in bs["equity_lines"] if l["system_key"] == "CURRENT_EARNINGS"][0]
    assert earnings["amount"] == D("1000.00")  # July sale excluded


def test_build_xlsx_response_roundtrip():
    import io
    import openpyxl
    from app.core.xlsx import Sheet, build_xlsx_bytes, build_xlsx_response

    sheets = [Sheet(name="P&L", headers=["Account", "Amount"],
                    rows=[["Sales Revenue", D("1000.00")], ["Net profit", D("250.00")]],
                    title="Income Statement")]

    resp = build_xlsx_response(sheets, filename="income-statement")
    assert resp.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "income-statement.xlsx" in resp.headers["content-disposition"]

    body = build_xlsx_bytes(sheets)
    assert body and body[:2] == b"PK"  # xlsx is a zip
    wb = openpyxl.load_workbook(io.BytesIO(body))
    assert wb.sheetnames == ["P&L"]
    ws = wb["P&L"]
    # Title row, then header row, then data
    assert ws.cell(row=1, column=1).value == "Income Statement"
    assert ws.cell(row=2, column=1).value == "Account"
    assert ws.cell(row=3, column=1).value == "Sales Revenue"
    assert ws.cell(row=3, column=2).value == 1000.0


def test_sheet_builders_shape():
    from app.api.statements import _pnl_sheets, _bs_sheets, _cf_sheets, _acct_sheets

    pnl = {"start": date(2026, 6, 1), "end": date(2026, 6, 30),
           "revenue_lines": [{"code": "4000", "name": "Sales", "system_key": "SALES_REVENUE", "amount": D("1000")}],
           "cogs_lines": [], "opex_lines": [],
           "revenue": D("1000"), "cogs": D("0"), "gross_profit": D("1000"),
           "operating_expenses": D("0"), "net_profit": D("1000")}
    sheets = _pnl_sheets(pnl)
    assert sheets and sheets[0].headers == ["Account", "Code", "Amount"]
    assert any("Net profit" in str(r[0]) for r in sheets[0].rows)

    bs = {"as_of": date(2026, 6, 30), "all_current": True,
          "asset_lines": [], "liability_lines": [], "equity_lines": [],
          "total_assets": D("0"), "total_liabilities": D("0"), "total_equity": D("0"),
          "balanced": True, "metal_position": [{"karat": "K21", "net_grams": D("100")}]}
    assert _bs_sheets(bs)  # at least one sheet
    cf = {"start": date(2026, 6, 1), "end": date(2026, 6, 30), "opening_cash": D("0"),
          "closing_cash": D("0"), "categories": [], "net_change": D("0"), "reconciles": True}
    assert _cf_sheets(cf)
    st = {"code": "1000", "name": "Cash", "opening_balance": D("0"), "closing_balance": D("0"),
          "rows": [], "start": date(2026, 6, 1), "end": date(2026, 6, 30)}
    assert _acct_sheets(st)
