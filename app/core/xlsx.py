"""Reusable xlsx export: tabular report → openpyxl workbook → StreamingResponse.

Mirrors the existing CSV StreamingResponse pattern (e.g. app/api/orders.py).
One workbook may hold multiple sheets (used by multi-section statements and the
KPI workbook in M7b).
"""
import io
from dataclasses import dataclass
from decimal import Decimal

from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass
class Sheet:
    name: str
    headers: list
    rows: list
    title: str | None = None


def _cell(v):
    if isinstance(v, Decimal):
        return float(v)
    return v


def _sanitize(name: str) -> str:
    # Excel sheet names: <=31 chars, none of : \ / ? * [ ]
    for ch in ':\\/?*[]':
        name = name.replace(ch, "-")
    return name[:31] or "Sheet"


def build_xlsx_bytes(sheets: list[Sheet]) -> bytes:
    """Render the sheets to an in-memory .xlsx and return its raw bytes.

    Kept separate from the response wrapper so it is directly unit-testable
    (re-openable with openpyxl.load_workbook) without consuming an async
    StreamingResponse body."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    bold = Font(bold=True)
    for sheet in sheets:
        ws = wb.create_sheet(title=_sanitize(sheet.name))
        if sheet.title:
            ws.append([sheet.title])
            ws.cell(row=ws.max_row, column=1).font = bold
        ws.append(list(sheet.headers))
        for c in range(1, len(sheet.headers) + 1):
            ws.cell(row=ws.max_row, column=c).font = bold
        for row in sheet.rows:
            ws.append([_cell(v) for v in row])
        # Auto-size columns (capped).
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 2, 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_xlsx_response(sheets: list[Sheet], *, filename: str) -> StreamingResponse:
    body = build_xlsx_bytes(sheets)
    return StreamingResponse(
        io.BytesIO(body), media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
